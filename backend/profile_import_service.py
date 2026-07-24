"""客户画像-文件清单导入编排。

流程:解析文件来源(现阶段 Excel,后期业务方接口) -> 落 customer_files
-> 逐文件取 OCR(先复用 archive_detect,没有再下载+OCR) -> 分类筛 4 类
-> 按 doc_extract_rules 提取 -> 归因写入客户档案。详见 docs/09。

本文件当前包含:文件来源协议 + Excel 解析(纯函数,可单测)。
run_import 编排逻辑在 slice (c) 落地。
"""
from collections import Counter
from typing import Optional, Protocol, TypedDict

from openpyxl import load_workbook


class ManifestFile(TypedDict):
    file_code: str
    filename: str
    folder_name: Optional[str]
    rel_path: Optional[str]
    client_name: Optional[str]  # 行级客户姓名(可空)


class FileManifest(TypedDict):
    client_name: str            # 主客户(客户姓名列众数)
    files: list[ManifestFile]


class FileSourceProvider(Protocol):
    """文件来源协议:给个来源描述,返回统一客户文件清单。

    Excel 是第一个实现(parse_excel_manifest);后期业务方查询接口实现
    同一协议(按 client_code/姓名远程拉清单),run_import 及下游零改动。
    """
    async def fetch_manifest(self, source: dict) -> FileManifest: ...


# 列别名映射:兼容原文件错别列名"文件啊名称"
_COLUMN_ALIASES = {
    "folder_name": ["售后文件夹名称", "文件夹名称"],
    "file_code": ["文件编码"],
    "client_name": ["客户姓名"],
    "filename": ["文件名称", "文件啊名称"],
    "file_path": ["文件路径"],
    "rel_path": ["相对路径"],
}
_REQUIRED_COLUMNS = ("file_code", "client_name", "filename")


def _cell_str(v) -> str:
    """单元格规整:None→'';整数型 float→整数字符串(防文件编码被读成数值);其余 strip。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def parse_excel_manifest(path: str) -> dict:
    """解析客户文件清单 Excel -> {client_name, files, skipped_rows, duplicates}。

    表头在第一行;必需列:文件编码/客户姓名/文件名称(错名列"文件啊名称"兼容)。
    file_code 与 filename 皆空的行跳过;file_code 重复保留首次。
    主客户取客户姓名列众数(同票取先出现者)。
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        header = None
        for row in rows:
            header = [_cell_str(c) for c in row]
            break
        if not header:
            raise ValueError("Excel 为空或缺少表头行")

        # 列名 -> 列索引
        col_idx: dict[str, int] = {}
        for field, aliases in _COLUMN_ALIASES.items():
            for alias in aliases:
                if alias in header:
                    col_idx[field] = header.index(alias)
                    break
        missing = [f for f in _REQUIRED_COLUMNS if f not in col_idx]
        if missing:
            names = "、".join("/".join(_COLUMN_ALIASES[f]) for f in missing)
            raise ValueError(f"Excel 缺少必需列: {names}")

        def _get(row, field) -> str:
            idx = col_idx.get(field)
            if idx is None or idx >= len(row):
                return ""
            return _cell_str(row[idx])

        files: list[ManifestFile] = []
        seen_codes: set[str] = set()
        name_counter: Counter = Counter()
        first_seen_name: dict[str, int] = {}
        skipped_rows = 0
        duplicates = 0

        for row in rows:
            file_code = _get(row, "file_code")
            filename = _get(row, "filename")
            client_name = _get(row, "client_name")
            if not file_code and not filename:
                skipped_rows += 1
                continue
            if client_name:
                if client_name not in first_seen_name:
                    first_seen_name[client_name] = len(name_counter)
                name_counter[client_name] += 1
            if file_code:
                if file_code in seen_codes:
                    duplicates += 1
                    continue
                seen_codes.add(file_code)
            files.append({
                "file_code": file_code,
                "filename": filename,
                "folder_name": _get(row, "folder_name") or None,
                "rel_path": _get(row, "rel_path") or None,
                "client_name": client_name or None,
            })

        if not files:
            raise ValueError("Excel 无有效数据行")
        if not name_counter:
            raise ValueError("客户姓名列无有效值,无法识别主客户")

        # 众数;同票取先出现者(Counter.most_common 同票保序依赖插入序,显式排)
        top_count = max(name_counter.values())
        client_name = min(
            (n for n, c in name_counter.items() if c == top_count),
            key=lambda n: first_seen_name[n],
        )
        return {
            "client_name": client_name,
            "files": files,
            "skipped_rows": skipped_rows,
            "duplicates": duplicates,
        }
    finally:
        wb.close()


# ==================== 导入编排 ====================

import asyncio
import os
import time
from datetime import datetime, timedelta

import doc_type_matcher
import event_service
import file_fetcher
import llm_service
import review_service
import text_extractor
from db import customer_file_crud, doc_extract_crud, profile_crud

# 4 类证件 -> 任务计数器字段名
_TYPE_COUNTER = {
    "id_card": "id_card_count",
    "hukou": "hukou_count",
    "degree_cert": "degree_cert_count",
    "birth_cert": "birth_cert_count",
}

# 原件落盘目录(相对 output/ 存 DB,绝对路径落盘);GC 按 file_keep_until 清理
_OUTPUT_DIR = os.path.normpath(os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "output"))
_CUSTOMER_FILES_SUBDIR = "customer_files"
_FILE_KEEP_DAYS = 30


def _persist_local_file(src_path: str, file_code: str, filename: str) -> str:
    """把下载的原件从临时目录移存到 output/customer_files/,返回相对 output/ 的路径。"""
    dest_dir = os.path.join(_OUTPUT_DIR, _CUSTOMER_FILES_SUBDIR)
    os.makedirs(dest_dir, exist_ok=True)
    safe_name = file_fetcher._sanitize_filename(filename or "file")
    dest = os.path.join(dest_dir, f"{file_code}_{safe_name}")
    os.replace(src_path, dest)
    return f"{_CUSTOMER_FILES_SUBDIR}/{file_code}_{safe_name}"


async def ensure_local_file(row: dict) -> tuple[str, str]:
    """返回 (原件绝对路径, mime)。本地有直接用;没有/已 GC 则刷新 URL 重下并顺延 30 天。

    抛 FileNotFoundError: 无本地原件且无 file_code 可重下。
    """
    rel = row.get("local_path")
    if rel:
        abs_path = os.path.join(_OUTPUT_DIR, rel)
        if os.path.exists(abs_path):
            return abs_path, row.get("mime_type") or "application/octet-stream"
    file_code = (row.get("file_code") or "").strip()
    if not file_code or file_code.startswith("nocode-"):
        raise FileNotFoundError("无本地原件且无法重新下载(缺文件编码)")
    url, _ = await file_fetcher.refresh_download_url(file_code)
    tmp_path, fname, mime = await file_fetcher.fetch_url_to_temp(url)
    rel_path = _persist_local_file(tmp_path, file_code, fname)
    await customer_file_crud.update_file_local(
        row["id"], local_path=rel_path,
        file_keep_until=datetime.now() + timedelta(days=_FILE_KEEP_DAYS))
    return os.path.join(_OUTPUT_DIR, rel_path), (mime or "application/octet-stream")


async def run_import(task_id: int) -> None:
    """导入任务主流程(主进程 asyncio.create_task 串行跑)。

    逐文件: 取 OCR(relink/复用/下载) -> 分类(关键词->LLM 兜底) -> 4 类提取+归因写库。
    单文件异常只标该文件 error,不杀任务;任务级异常落 finish_import_task('error')。
    """
    task = await customer_file_crud.get_import_task(task_id)
    if not task:
        return
    counters = {
        "processed_files": 0, "reused_count": 0, "relinked_count": 0,
        "fresh_ocr_count": 0, "failed_count": 0, "extracted_count": 0,
        "id_card_count": 0, "hukou_count": 0, "degree_cert_count": 0,
        "birth_cert_count": 0, "needs_review_count": 0,
    }
    try:
        rows = await customer_file_crud.list_pending_files(task_id)
        for row in rows:
            counters["processed_files"] += 1
            await customer_file_crud.update_task_progress(
                task_id, current_file=row.get("filename") or row.get("file_code"), **counters)
            try:
                await _process_one_file(task, row, counters)
            except Exception as e:
                counters["failed_count"] += 1
                await customer_file_crud.mark_file_error(row["id"], f"{type(e).__name__}: {e}")
                print(f"[profile_import:{task_id}] 文件 {row.get('file_code')} 处理失败: {e}")
            await customer_file_crud.update_task_progress(task_id, **counters)

        await customer_file_crud.finish_import_task(task_id, "done")
        event_service.log_event(
            event_service.INFO, event_service.CATEGORY_PROFILE_IMPORT_DONE,
            f"客户画像导入完成: {task.get('client_name')} 共 {counters['processed_files']} 个文件",
            context={"task_id": task_id, **counters},
        )
    except Exception as e:
        await customer_file_crud.finish_import_task(task_id, "error", f"{type(e).__name__}: {e}")
        event_service.log_event(
            event_service.ERROR, event_service.CATEGORY_PROFILE_IMPORT_ERROR,
            f"客户画像导入失败: {e}",
            context={"task_id": task_id, "error": str(e)},
        )


async def _process_one_file(task: dict, row: dict, counters: dict) -> None:
    task_id = task["id"]
    file_code = row.get("file_code") or ""
    ocr_text = row.get("ocr_text") or ""

    # ---- 1) 取 OCR ----
    if row.get("status") == "done" and ocr_text:
        counters["relinked_count"] += 1  # 本库已有 done 行,直接复用
    else:
        await customer_file_crud.set_file_status(row["id"], "fetching")
        reused = await customer_file_crud.find_reusable_ocr(file_code)
        if reused:
            ocr_text = reused["ocr_text"] or ""
            await customer_file_crud.update_file_ocr(
                row["id"], status="ocr", ocr_source="reused", ocr_text=ocr_text,
                mime_type=reused.get("mime_type"),
                page_count=reused.get("page_count"),
                char_count=reused.get("char_count"),
            )
            counters["reused_count"] += 1
        else:
            if file_code.startswith("nocode-"):
                raise ValueError("缺少文件编码,无法刷新下载地址")
            url, _ = await file_fetcher.refresh_download_url(file_code)
            tmp_path, fname, mime = await file_fetcher.fetch_url_to_temp(url)
            result = await text_extractor.extract_text(tmp_path, mime)
            ocr_text = result.get("text") or ""
            # 原件落盘留存(30 天 GC;复核在线查看用),不再删除
            rel_path = _persist_local_file(tmp_path, file_code, fname)
            await customer_file_crud.update_file_local(
                row["id"], local_path=rel_path,
                file_keep_until=datetime.now() + timedelta(days=_FILE_KEEP_DAYS))
            await customer_file_crud.update_file_ocr(
                row["id"], status="ocr", ocr_source="fresh", ocr_text=ocr_text,
                mime_type=mime,
                page_count=result.get("page_count"),
                char_count=result.get("char_count"),
            )
            counters["fresh_ocr_count"] += 1

    # ---- 2) 分类 ----
    doc_type = None
    classify_by = "none"
    classify_score = None
    if ocr_text.strip():
        m = doc_type_matcher.classify(
            row.get("folder_name"), row.get("filename"),
            ocr_text[: doc_type_matcher.OCR_HEAD_CHARS], row.get("rel_path"))
        if m["doc_type"]:
            doc_type = m["doc_type"]
            classify_by, classify_score = "keyword", m["score"]
        else:
            r = await asyncio.to_thread(
                llm_service.recognize_doc_type, ocr_text[:2000],
                task_id=str(task_id), file_id=file_code)
            doc_type = r["doc_type"]
            classify_by, classify_score = "llm", r["confidence"]
        await customer_file_crud.update_file_classify(
            row["id"], doc_type=doc_type, classify_by=classify_by,
            classify_score=classify_score)
        if doc_type in _TYPE_COUNTER:
            counters[_TYPE_COUNTER[doc_type]] += 1
    else:
        await customer_file_crud.update_file_classify(
            row["id"], doc_type=None, classify_by="none", classify_score=None)

    # ---- 3) 4 类提取 + 归因写库(无 active 规则则留 skipped 痕) ----
    extract_outcome = None
    if doc_type in llm_service.DOC_EXTRACT_TYPES and ocr_text.strip():
        extract_outcome = await _extract_one(task, row, doc_type, ocr_text, counters)

    # ---- 4) 质量评级(纯规则,复核轻重缓急的依据) ----
    q = review_service.evaluate_file_quality(
        ocr_text=ocr_text, folder_name=row.get("folder_name"),
        doc_type=doc_type, classify_by=classify_by, classify_score=classify_score,
        extract_status=(extract_outcome or {}).get("status"),
        extract_skip_reason=(extract_outcome or {}).get("skip_reason"),
        id_masked=bool((extract_outcome or {}).get("id_masked")))
    await customer_file_crud.update_file_review(
        row["id"], quality_score=q["quality_score"],
        review_status=q["review_status"], review_reason=q["review_reason"])
    if q["review_status"] == "needs_review":
        counters["needs_review_count"] += 1


async def _extract_one(task: dict, row: dict, doc_type: str,
                       ocr_text: str, counters: dict) -> dict:
    """4 类证件提取 + 归因 + 写 profile_person_fields。返回 outcome 供质量评级使用。"""
    task_id = task["id"]
    household_id = task.get("household_id")
    file_code = row.get("file_code") or ""
    t0 = time.time()
    outcome = {"status": None, "skip_reason": None, "id_masked": False}

    rule = await doc_extract_crud.get_active_rule(doc_type)
    if not rule:
        await doc_extract_crud.insert_result(
            customer_file_id=row["id"], import_task_id=task_id, file_id=file_code,
            client_id=task.get("client_id"), doc_type=doc_type, rule_id=None,
            rule_version=None, status="skipped", skip_reason="no_active_rule")
        event_service.log_event(
            event_service.INFO, event_service.CATEGORY_EXTRACT_SKIP,
            f"证件类型 {doc_type} 无 active 提取规则,跳过提取",
            context={"task_id": task_id, "file_code": file_code, "doc_type": doc_type,
                     "reason": "no_active_rule"})
        outcome.update(status="skipped", skip_reason="no_active_rule")
        return outcome

    # 乱码文本不提取(垃圾进垃圾出:实测歪扫户口页会产出乱码人名并错误建人;
    # 该文件已在质量评级中标 garbled 进复核队列)
    if review_service.is_garbled(ocr_text):
        await doc_extract_crud.insert_result(
            customer_file_id=row["id"], import_task_id=task_id, file_id=file_code,
            client_id=task.get("client_id"), doc_type=doc_type,
            rule_id=rule["id"], rule_version=rule["version"],
            status="skipped", skip_reason="garbled_text")
        event_service.log_event(
            event_service.WARN, event_service.CATEGORY_EXTRACT_SKIP,
            f"OCR 乱码,跳过提取: {file_code}",
            context={"task_id": task_id, "file_code": file_code, "doc_type": doc_type,
                     "reason": "garbled_text"})
        outcome.update(status="skipped", skip_reason="garbled_text")
        return outcome

    try:
        raw = await asyncio.to_thread(
            llm_service.extract_doc_fields, ocr_text, rule,
            task_id=str(task_id), file_id=file_code)
        extracted = raw.get("fields") or {}

        # 字段清洗:去空;masked 不候选归因/写库(在 apply 内记 skipped_masked)
        field_items = []
        for f in rule.get("fields") or []:
            v = extracted.get(f.get("key"))
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            field_items.append({
                "key": f.get("key"), "label": f.get("label"),
                "value": str(v).strip(),
                "column": (f.get("target") or {}).get("column"),
                "layer": (f.get("target") or {}).get("layer"),
                "entity": (f.get("target") or {}).get("entity") or "person",
            })

        id_number = next(
            (it["value"] for it in field_items
             if it["column"] == "id_number"
             and not doc_extract_crud.is_masked(it["value"])
             and doc_extract_crud.valid_id_number(it["value"])), None)
        name = next(
            (it["value"] for it in field_items
             if it["column"] == "name" and not doc_extract_crud.is_masked(it["value"])), None)
        if name and not profile_crud.plausible_person_name(name):
            name = None  # 乱码假名(如 "钅 lil蝴哪")不参与归因/建人,按无姓名走 no_person
        # 拼音名归因(批复/永居卡等英文证件没有中文名;词序无关匹配 person_fields.name_en)
        name_en = next(
            (it["value"] for it in field_items
             if it["column"] == "name_en" and not doc_extract_crud.is_masked(it["value"])), None)
        outcome["id_masked"] = any(
            it["column"] == "id_number" and doc_extract_crud.is_masked(it["value"])
            for it in field_items)

        match = await profile_crud.find_person_match(household_id, id_number, name, name_en)
        case_items = [it for it in field_items if it.get("entity") == "case"]
        if match.get("person_id") is None and not name and not case_items:
            await doc_extract_crud.insert_result(
                customer_file_id=row["id"], import_task_id=task_id, file_id=file_code,
                client_id=task.get("client_id"), doc_type=doc_type,
                rule_id=rule["id"], rule_version=rule["version"],
                status="skipped", skip_reason="no_person",
                extracted=extracted,
                elapsed_ms=int((time.time() - t0) * 1000))
            event_service.log_event(
                event_service.WARN, event_service.CATEGORY_EXTRACT_SKIP,
                f"提取结果无姓名无法归属: {file_code}",
                context={"task_id": task_id, "file_code": file_code, "doc_type": doc_type,
                         "reason": "no_person"})
            outcome.update(status="skipped", skip_reason="no_person")
            return outcome

        write = await profile_crud.apply_extracted_fields_v2(
            household_id, match, field_items, source_file_id=row["id"])
        # entity=asset 的字段写入家庭资产表(房产证类),去重靠 attrs.cert_no/name
        asset_items = [it for it in field_items if it.get("entity") == "asset"]
        if asset_items:
            aw = await profile_crud.apply_extracted_asset(
                household_id, write.get("person_id"), asset_items,
                source_file_id=row["id"])
            write["mapped"] += aw["mapped"]
            write["write_stats"].update(aw["stats"])
        # entity=case 的字段写入案件时间线(递交/签收/批复里程碑,家庭单案件)
        if case_items:
            cw = await profile_crud.apply_case_milestones(
                household_id, case_items, source_file_id=row["id"])
            write["mapped"] += cw["mapped"]
            write["write_stats"].update(cw["stats"])
        await doc_extract_crud.insert_result(
            customer_file_id=row["id"], import_task_id=task_id, file_id=file_code,
            client_id=task.get("client_id"), doc_type=doc_type,
            rule_id=rule["id"], rule_version=rule["version"],
            status="done", extracted=extracted,
            mapped=write["mapped"], write_stats=write["write_stats"],
            elapsed_ms=int((time.time() - t0) * 1000))
        counters["extracted_count"] += 1
        event_service.log_event(
            event_service.INFO, event_service.CATEGORY_EXTRACT_DONE,
            f"证件信息提取完成: {file_code} ({doc_type})",
            context={"task_id": task_id, "file_code": file_code, "doc_type": doc_type,
                     "rule_version": rule["version"],
                     "matched_by": write["write_stats"].get("matched_by"),
                     "writes": {k: v for k, v in write["write_stats"].items()
                                if k in ("written", "updated", "person_created")}})
        outcome.update(status="done")
        return outcome
    except Exception as e:
        await doc_extract_crud.insert_result(
            customer_file_id=row["id"], import_task_id=task_id, file_id=file_code,
            client_id=task.get("client_id"), doc_type=doc_type,
            rule_id=rule["id"], rule_version=rule["version"],
            status="error", error_msg=f"{type(e).__name__}: {e}",
            elapsed_ms=int((time.time() - t0) * 1000))
        event_service.log_event(
            event_service.WARN, event_service.CATEGORY_EXTRACT_ERROR,
            f"证件信息提取失败: {file_code}: {e}",
            context={"task_id": task_id, "file_code": file_code, "doc_type": doc_type,
                     "error": str(e)})
        outcome.update(status="error")
        return outcome
